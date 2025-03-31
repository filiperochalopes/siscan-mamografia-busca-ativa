from flask import Flask, render_template, request, redirect, url_for, session, send_file, flash
from dotenv import load_dotenv
import os, re, uuid, secrets
from io import BytesIO
from pathlib import Path
from datetime import datetime
from src.extrator_laudo.mamografia import SiscanReportMammographyExtract
import pandas as pd
import numpy as np
from openpyxl import load_workbook


load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "minha_chave_secreta")  # Utilize uma chave segura em produção
TOKEN = os.getenv("TOKEN")

@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        token = request.form.get("token")
        if token == TOKEN:
            session["logged_in"] = True
            return redirect(url_for("upload"))
        else:
            flash("Token incorreto!")
    return render_template("login.html.j2")

@app.route("/upload", methods=["GET", "POST"])
def upload():
    if not session.get("logged_in"):
        return redirect(url_for("login"))

    download_url = None

    if request.method == "POST":
        file = request.files.get("file")
        if file and file.filename != "":
            timestamp = datetime.now().strftime("%Y%m%d%H%M")
            nome_arquivo = f"resultado_processamento_laudos_siscan_{timestamp}.xlsx"
            export_dir = Path("src/static/exports")
            export_dir.mkdir(parents=True, exist_ok=True)

            # Salva arquivo e processa
            temp_pdf = Path("tmp/input.pdf")
            Path("tmp").mkdir(exist_ok=True)
            file.save(temp_pdf)

            extrator = SiscanReportMammographyExtract("tmp", "tmp")
            _, df = extrator.process()
            print(df.columns, flush=True)
            print(df.head(), flush=True)

            # Extrai o número do BIRADS a partir de uma string como "Categoria 2 - achados mamográficos benignos"
            def extrair_birads(texto):
                try:
                    # Divide a string e pega o segundo elemento (número)
                    numero = int(texto.split()[1])
                    return numero
                except Exception:
                    return np.nan

            # Calcula a idade a partir da data de nascimento
            def calcular_idade(data_nascimento):
                if pd.isna(data_nascimento):
                    return np.nan
                hoje = datetime.today().date()
                try:
                    # Se a data vem como string, converte para datetime (assumindo formato dia/mês/ano)
                    if isinstance(data_nascimento, str):
                        data_nascimento = pd.to_datetime(data_nascimento, dayfirst=True).date()
                    else:
                        data_nascimento = data_nascimento.date()
                    idade = hoje.year - data_nascimento.year - ((hoje.month, hoje.day) < (data_nascimento.month, data_nascimento.day))
                    return idade
                except Exception:
                    return np.nan

            # Verifica se a mama é densa: retorna 1 se o valor for "Predominantemente Densa" ou "Densa"
            def is_densa(tipo_de_mama):
                return 1 if tipo_de_mama in ["Predominantemente Densa", "Densa"] else 0

            # Verifica se há indicação para USG, usando vários campos
            def verificar_usg(row):
                # Se alguma mama for densa
                if is_densa(row.get("resultado_exame__mama_direita__tipo_de_mama", "")) or is_densa(row.get("resultado_exame__mama_esquerda__tipo_de_mama", "")):
                    return 1
                # Se na classificação radiológica aparecer "Categoria 0"
                if ("Categoria 0" in str(row.get("resultado_exame__classificacao_radiologica__mama_direita", "")) or 
                    "Categoria 0" in str(row.get("resultado_exame__classificacao_radiologica__mama_esquerda", ""))):
                    return 1
                # Se em recomendações constar "Complemento com ultrassonografia"
                if "Complemento com ultrassonografia" in str(row.get("resultado_exame__recomendacoes", "")):
                    return 1
                return 0

            # Define a coluna "Alterado" com base no BIRADS e se há indicação para USG
            def definir_alterado(row, birads_md, birads_me, usg):
                # Se qualquer lado apresentar BIRADS 4 ou superior, marca 2
                if (not np.isnan(birads_md) and birads_md >= 4) or (not np.isnan(birads_me) and birads_me >= 4):
                    return 2
                # Se qualquer lado tiver BIRADS 3 ou se houver indicação de USG, marca 1
                if (birads_md == 3 or birads_me == 3) or usg == 1:
                    return 1
                # Caso contrário, marca 0 (BIRADS 1 ou 2)
                return 0

            # === Etapa 1: Gerar a pasta de processamento e o token ===

            # Gera um timestamp no formato yyyymmddhhii
            timestamp = datetime.now().strftime("%Y%m%d%H%M")
            # Gera um hash de 12 caracteres (6 bytes em hexadecimal)
            folder_hash = secrets.token_hex(6)
            # Nome da pasta: ex: 202303311530_ab12cd34ef56
            folder_name = f"{timestamp}_{folder_hash}"

            # Define a pasta de destino dentro da pasta static (por exemplo: src/static)
            static_folder = Path("src/static")
            dest_folder = static_folder / folder_name
            dest_folder.mkdir(parents=True, exist_ok=True)

            # Gera o token de acesso (64 bits, ou seja, 8 bytes em hexadecimal – 16 caracteres)
            token = secrets.token_hex(8)

            # === Etapa 3: Renomear e mover os PDFs para a nova pasta ===
            # Supondo que os PDFs gerados inicialmente tenham nome padrão: laudo_<pagina>.pdf
            # e estejam em "tmp/pages". Montamos um dicionário para mapear número da página ao novo nome UUID.

            tmp_pages = Path("tmp/pages")
            mapping_pdf = {}
            for pdf in tmp_pages.glob("*.pdf"):
                # Procura um padrão que termine com _page_<número>
                match = re.search(r'_page_(\d+)$', pdf.stem)
                if match:
                    page_number = int(match.group(1))
                    new_name = f"{uuid.uuid4()}.pdf"
                    dest_path = dest_folder / new_name
                    pdf.rename(dest_path)
                    mapping_pdf[page_number] = new_name
                else:
                    # Se o padrão não for encontrado, pode registrar um aviso ou ignorar o arquivo
                    continue

            # --- Transformação do DataFrame ---

            novo_df = pd.DataFrame()

            # Mapeia as colunas diretamente ou com conversões:
            novo_df["Nome"] = df["paciente__nome"]
            # Converte a coluna de data de nascimento, assumindo que vem no formato "15/01/1976"
            data_nasc_dt = pd.to_datetime(df["paciente__data_do_nascimento"], format="%d/%m/%Y", errors="coerce")

            # Para calcular a idade, mantenha o objeto datetime
            novo_df["Idade"] = data_nasc_dt.apply(calcular_idade)

            # Se preferir salvar a data formatada como mm/dd/yyyy (apenas para exibição no Excel)
            novo_df["Data de nascimento"] = data_nasc_dt.dt.strftime("%m/%d/%Y")
            novo_df["Nome da mãe"] = df["paciente__mae"]
            # CNS: assumindo que seja o número do cartão de saúde disponível em "paciente__cartao_sus"
            novo_df["CNS"] = df["paciente__cartao_sus"]

            # Data do exame – pode vir de "prestador_de_servico__data_da_realizacao" (ou outra, conforme seu contexto)
            # Extrai "Data do exame" a partir da coluna "geral__emissao" (formato dd/mm/yyyy)
            novo_df["Data do exame"] = pd.to_datetime(df["geral__emissao"], format="%d/%m/%Y", errors="coerce").dt.strftime("%m/%d/%Y")

            novo_df["Unidade de saúde"] = df["unidade_de_saude__nome"]
            novo_df["CNES"] = df["unidade_de_saude__cnes"]
            # Tipo do Exame: mantém o valor de "resultado_exame__indicacao__tipo_de_mamografia"
            novo_df["Tipo do Exame"] = df["resultado_exame__indicacao__tipo_de_mamografia"]

            # BIRADS: extraindo os números das colunas de classificação radiológica
            # Extração dos BIRADS já feita:
            birads_md = df["resultado_exame__classificacao_radiologica__mama_direita"].apply(lambda x: extrair_birads(str(x)))
            birads_me = df["resultado_exame__classificacao_radiologica__mama_esquerda"].apply(lambda x: extrair_birads(str(x)))

            # Converter para string com o prefixo "BI-RADS" e garantir que seja um inteiro (se existir)
            novo_df["MD - BIRADS"] = birads_md.apply(lambda x: f"BI-RADS {int(x)}" if pd.notna(x) else "")
            novo_df["ME - BIRADS"] = birads_me.apply(lambda x: f"BI-RADS {int(x)}" if pd.notna(x) else "")

            # Mama densa: usando os campos "tipo_de_mama"
            novo_df["MD - Mama densa"] = df["resultado_exame__mama_direita__tipo_de_mama"].apply(is_densa)
            novo_df["ME - Mama densa"] = df["resultado_exame__mama_esquerda__tipo_de_mama"].apply(is_densa)

            # USG: coluna booleana baseada em várias condições
            novo_df["USG"] = df.apply(verificar_usg, axis=1)

            # Alterado: baseado nos BIRADS e indicação de USG
            novo_df["Alterado"] = df.apply(lambda row: definir_alterado(
                row,
                extrair_birads(str(row.get("resultado_exame__classificacao_radiologica__mama_direita", ""))),
                extrair_birads(str(row.get("resultado_exame__classificacao_radiologica__mama_esquerda", ""))),
                verificar_usg(row)
            ), axis=1)

            # Link para arquivo: usando o número da página presente em "geral__pagina"
            # Atualiza o link para arquivo usando o mapeamento criado.
            # A partir do valor de "geral__pagina", buscamos o novo nome do arquivo e montamos o link absoluto
            APP_URL = "https://siscan.filipelopes.med.br"  # Substitua pelo domínio real da sua aplicação
            novo_df["Link para arquivo"] = df["geral__pagina"].apply(
                lambda x: f"{APP_URL}/static/{folder_name}/{mapping_pdf.get(int(x), '')}?token={token}"
                if pd.notna(x) and int(x) in mapping_pdf else ""
)

            # Colunas de controle operacional (inicialmente vazias ou default)
            novo_df["Pendente"] = 1
            novo_df["Data de ação"] = pd.NaT
            novo_df["Resultado da ação"] = ""
            novo_df["Observações"] = ""

            # Reordena as colunas conforme a especificação
            colunas_ordenadas = [
                "Nome", "Data de nascimento", "Idade", "Nome da mãe", "CNS",
                "Unidade de saúde", "CNES", "Data do exame", "Tipo do Exame",
                "MD - BIRADS", "MD - Mama densa", "ME - BIRADS", "ME - Mama densa",
                "Alterado", "USG", "Link para arquivo", "Pendente",
                "Data de ação", "Resultado da ação", "Observações"
            ]
            novo_df = novo_df[colunas_ordenadas]

            # Ordena as linhas: Alterado DESC, depois Tipo do Exame, CNES, Unidade de saúde e Nome
            novo_df.sort_values(
                by=["Alterado", "Tipo do Exame", "CNES", "Unidade de saúde", "Nome"],
                ascending=[False, True, True, True, True],
                inplace=True
            )

            # Salva o DataFrame em Excel
            caminho_excel = export_dir / nome_arquivo
            novo_df.to_excel(caminho_excel, index=False)

            # Abre o arquivo Excel para ajustar os hiperlinks
            wb = load_workbook(caminho_excel)
            ws = wb.active

            # Supondo que a coluna "Link para arquivo" esteja na posição determinada pela ordem das colunas
            link_col_index = colunas_ordenadas.index("Link para arquivo") + 1  # índice 1-base

            # Itera nas linhas (pulando o cabeçalho na linha 1)
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=link_col_index)
                if cell.value:
                    # Armazena o link original
                    link = cell.value
                    # Define o hiperlink com o link original
                    cell.hyperlink = link
                    # Define o texto que será exibido na célula
                    cell.value = "Acessar o laudo"
                    # Aplica o estilo de hiperlink (geralmente azul e sublinhado)
                    cell.style = "Hyperlink"

            # Salva as alterações no arquivo Excel
            wb.save(caminho_excel)

            download_url = url_for('static', filename=f'exports/{nome_arquivo}')

    return render_template("upload.html.j2", download_url=download_url or "")

@app.route("/download")
def download():
    if not session.get("logged_in"):
        return redirect(url_for("login"))
    # Cria um arquivo CSV dummy para download (substitua pelo seu processamento)
    output = BytesIO()
    output.write(b"col1,col2,col3\n1,2,3\n4,5,6")
    output.seek(0)
    return send_file(output, mimetype="text/csv", as_attachment=True, attachment_filename="planilha.csv")
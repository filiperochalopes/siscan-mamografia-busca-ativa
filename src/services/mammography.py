from __future__ import annotations
from pathlib import Path
import pandas as pd
from src.utils.files import clear_tmp
from src.utils.mammography import (
    extrair_birads,
    calcular_idade,
    is_densa,
    verificar_usg,
    definir_alterado,
)
from src.extrator_laudo.mamografia import SiscanReportMammographyExtract
from datetime import datetime
import re
import uuid
import secrets
from openpyxl import load_workbook
from flask import url_for
from src.services.config import Config
from src.utils.validators import is_valid_siscan_df
from werkzeug.datastructures import FileStorage as File

class ReportService:
    def __init__(self, file: File):
        self.file = file
        self.extrator = SiscanReportMammographyExtract("tmp", "tmp")

    def process(self) -> pd.DataFrame:
        _, df = self.extrator.process()
        
        if not is_valid_siscan_df(df):
            raise ValueError("Erro: O arquivo enviado não é um laudo de mamografia SISCAN válido.")
    
        if df.empty or not df["paciente__nome"].any():
            raise ValueError("PDF não contém laudo válido")
        return df

    def build_excel(self) -> str:
        timestamp = datetime.now().strftime("%Y%m%d%H%M")
        nome_arquivo = f"resultado_processamento_laudos_siscan_{timestamp}.xlsx"
        export_dir = Path("src/static/exports")
        export_dir.mkdir(parents=True, exist_ok=True)

        # Salva arquivo e processa
        temp_pdf = Path("tmp/input.pdf")
        Path("tmp").mkdir(exist_ok=True)
        self.file.save(temp_pdf)

        try:
            df = self.process()

        except Exception as e:
            raise ValueError(f"Erro ao processar o PDF: {str(e)}")

        print(df.columns, flush=True)
        print(df.head(), flush=True)
        
        # === Etapa 1: Gerar a pasta de processamento e o token ===

        # Gera um timestamp no formato yyyymmddhhii
        timestamp = datetime.now().strftime("%Y%m%d%H%M")
        # Gera um hash de 12 caracteres (6 bytes em hexadecimal)
        folder_hash = secrets.token_hex(6)
        # Nome da pasta: ex: 202303311530_ab12cd34ef56
        folder_name = f"{timestamp}_{folder_hash}"

        # Define a pasta de destino dentro da pasta static (por exemplo: src/static)
        static_folder = Path("src/static")
        dest_folder = static_folder / folder_name
        dest_folder.mkdir(parents=True, exist_ok=True)

        # Gera o token de acesso (64 bits, ou seja, 8 bytes em hexadecimal – 16 caracteres)
        token = secrets.token_hex(8)

        # === Etapa 3: Renomear e mover os PDFs para a nova pasta ===
        # Supondo que os PDFs gerados inicialmente tenham nome padrão: laudo_<pagina>.pdf
        # e estejam em "tmp/pages". Montamos um dicionário para mapear número da página ao novo nome UUID.

        tmp_pages = Path("tmp/pages")
        mapping_pdf = {}
        for pdf in tmp_pages.glob("*.pdf"):
            # Procura um padrão que termine com _page_<número>
            match = re.search(r'_page_(\d+)$', pdf.stem)
            if match:
                page_number = int(match.group(1))
                new_name = f"{uuid.uuid4()}.pdf"
                dest_path = dest_folder / new_name
                pdf.rename(dest_path)
                mapping_pdf[page_number] = new_name
            else:
                # Se o padrão não for encontrado, pode registrar um aviso ou ignorar o arquivo
                continue

        # --- Transformação do DataFrame ---

        novo_df = pd.DataFrame()

        # Mapeia as colunas diretamente ou com conversões:
        novo_df["Nome"] = df["paciente__nome"]
        # Converte a coluna de data de nascimento, assumindo que vem no formato "15/01/1976"
        data_nasc_dt = pd.to_datetime(df["paciente__data_do_nascimento"], format="%d/%m/%Y", errors="coerce")

        # Para calcular a idade, mantenha o objeto datetime
        novo_df["Idade"] = data_nasc_dt.apply(calcular_idade)

        # Se preferir salvar a data formatada como mm/dd/yyyy (apenas para exibição no Excel)
        novo_df["Data de nascimento"] = data_nasc_dt.dt.strftime("%m/%d/%Y")
        novo_df["Nome da mãe"] = df["paciente__mae"]
        # CNS: assumindo que seja o número do cartão de saúde disponível em "paciente__cartao_sus"
        novo_df["CNS"] = df["paciente__cartao_sus"]

        # Data do exame – pode vir de "prestador_de_servico__data_da_realizacao" (ou outra, conforme seu contexto)
        # Extrai "Data do exame" a partir da coluna "geral__emissao" (formato dd/mm/yyyy)
        novo_df["Data do exame"] = pd.to_datetime(df["geral__emissao"], format="%d/%m/%Y", errors="coerce").dt.strftime("%m/%d/%Y")

        novo_df["Unidade de saúde"] = df["unidade_de_saude__nome"]
        novo_df["CNES"] = df["unidade_de_saude__cnes"]
        # Tipo do Exame: mantém o valor de "resultado_exame__indicacao__tipo_de_mamografia"
        novo_df["Tipo do Exame"] = df["resultado_exame__indicacao__tipo_de_mamografia"]

        # BIRADS: extraindo os números das colunas de classificação radiológica
        # Extração dos BIRADS já feita:
        birads_md = df["resultado_exame__classificacao_radiologica__mama_direita"].apply(lambda x: extrair_birads(str(x)))
        birads_me = df["resultado_exame__classificacao_radiologica__mama_esquerda"].apply(lambda x: extrair_birads(str(x)))

        # Converter para string com o prefixo "BI-RADS" e garantir que seja um inteiro (se existir)
        novo_df["MD - BIRADS"] = birads_md.apply(lambda x: f"BI-RADS {int(x)}" if pd.notna(x) else "")
        novo_df["ME - BIRADS"] = birads_me.apply(lambda x: f"BI-RADS {int(x)}" if pd.notna(x) else "")

        # Mama densa: usando os campos "tipo_de_mama"
        novo_df["MD - Mama densa"] = df["resultado_exame__mama_direita__tipo_de_mama"].apply(is_densa)
        novo_df["ME - Mama densa"] = df["resultado_exame__mama_esquerda__tipo_de_mama"].apply(is_densa)

        # USG: coluna booleana baseada em várias condições
        novo_df["USG"] = df.apply(verificar_usg, axis=1)

        # Alterado: baseado nos BIRADS e indicação de USG
        novo_df["Alterado"] = df.apply(lambda row: definir_alterado(
            row,
            extrair_birads(str(row.get("resultado_exame__classificacao_radiologica__mama_direita", ""))),
            extrair_birads(str(row.get("resultado_exame__classificacao_radiologica__mama_esquerda", ""))),
            verificar_usg(row)
        ), axis=1)

        # Link para arquivo: usando o número da página presente em "geral__pagina"
        # Atualiza o link para arquivo usando o mapeamento criado.
        # A partir do valor de "geral__pagina", buscamos o novo nome do arquivo e montamos o link absoluto
        novo_df["Link para arquivo"] = df["geral__pagina"].apply(
            lambda x: f"{Config.APP_URL}/static/{folder_name}/{mapping_pdf.get(int(x), '')}?token={token}"
            if pd.notna(x) and int(x) in mapping_pdf else ""
)

        # Colunas de controle operacional (inicialmente vazias ou default)
        novo_df["Pendente"] = 1
        novo_df["Data de ação"] = pd.NaT
        novo_df["Resultado da ação"] = ""
        novo_df["Observações"] = ""

        # Reordena as colunas conforme a especificação
        colunas_ordenadas = [
            "Nome", "Data de nascimento", "Idade", "Nome da mãe", "CNS",
            "Unidade de saúde", "CNES", "Data do exame", "Tipo do Exame",
            "MD - BIRADS", "MD - Mama densa", "ME - BIRADS", "ME - Mama densa",
            "Alterado", "USG", "Link para arquivo", "Pendente",
            "Data de ação", "Resultado da ação", "Observações"
        ]
        novo_df = novo_df[colunas_ordenadas]

        # Ordena as linhas: Alterado DESC, depois Tipo do Exame, CNES, Unidade de saúde e Nome
        novo_df.sort_values(
            by=["Alterado", "Tipo do Exame", "CNES", "Unidade de saúde", "Nome"],
            ascending=[False, True, True, True, True],
            inplace=True
        )

        # Salva o DataFrame em Excel
        caminho_excel = export_dir / nome_arquivo
        novo_df.to_excel(caminho_excel, index=False)

        # Abre o arquivo Excel para ajustar os hiperlinks
        wb = load_workbook(caminho_excel)
        ws = wb.active

        # Supondo que a coluna "Link para arquivo" esteja na posição determinada pela ordem das colunas
        link_col_index = colunas_ordenadas.index("Link para arquivo") + 1  # índice 1-base

        # Itera nas linhas (pulando o cabeçalho na linha 1)
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=link_col_index)
            if cell.value:
                # Armazena o link original
                link = cell.value
                # Define o hiperlink com o link original
                cell.hyperlink = link
                # Define o texto que será exibido na célula
                cell.value = "Acessar o laudo"
                # Aplica o estilo de hiperlink (geralmente azul e sublinhado)
                cell.style = "Hyperlink"

        # Salva as alterações no arquivo Excel
        wb.save(caminho_excel)
        clear_tmp()
        return url_for('static', filename=f'exports/{nome_arquivo}')